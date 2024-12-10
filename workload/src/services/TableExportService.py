from tempfile import NamedTemporaryFile

from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from sqlalchemy import select

from src.models import *
from src.utils import Database


class ExportService:
    def __init__(self, database: Database):
        self.db = database

    async def export_workload(self) -> str:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Лекциии"

        headers = ["Факультет", "Семестр", "Поток", "Наименование", "Преподаватель", "Нагрузка", "Тип", "Группы", ""]
        sheet.append(headers)

        async with self.db.session_factory() as session:
            containers = await session.execute(select(WorkloadContainer).distinct())
            containers: list[WorkloadContainer] = containers.scalars().unique().all()
            for container in containers:
                groups =  set()
                for workload in container.workloads:
                    groups.update([group.name for group in workload.groups])

                lessons: set[Lesson] = set()
                for workload in container.workloads:
                    lessons.add(workload.lesson)

                sheet.append([
                    ", ".join([str(lesson.faculty) for lesson in lessons]) if len(lessons) != 0 else "",
                    ", ".join([str(lesson.semester) for lesson in lessons]) if len(lessons) != 0 else "",
                    ", ".join([str(lesson.stream) for lesson in lessons]) if len(lessons) != 0 else "",
                    ", ".join([lesson.name for lesson in lessons]) if len(lessons) != 0 else "",
                    container.employee.name if container.employee is not None else "",
                    container.sum_workload,
                    container.workload_type if container.workload_type is not None else "",
                    ", ".join(groups) if len(groups) != 0 else "",
                    "Объединенный поток!!!" if len(lessons) != 1 else "",
                ])

        ExportService.set_width_and_alignment(sheet, [10, 10, 7, 40, 40, 10, 15, 10])
        temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        return temp_file.name

    async def export_employees(self) -> str:
        """
        :return: имя файла для экспорта
        """

        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Лекциии'

        headers = ['id', 'Имя', 'Номер', 'Почта', 'Пол', 'День рождения', 'Предпочитаемый факультет',
                   'Департамент', 'Должность', 'Ставка', 'Тип занятости', 'Дата завершения контракта']

        sheet.append(headers)

        async with self.db.session_factory() as session:
            employees_d = await session.execute(select(Employee).distinct())
            employees: list[Employee] = employees_d.scalars().unique().all()

            for employee in employees:
                for idx, position in enumerate(employee.positions):
                    sheet.append([
                        employee.id if idx == 0 else "",
                        employee.name if idx == 0 else "",
                        employee.phone if idx == 0 else "",
                        employee.mail if idx == 0 else "",
                        employee.gender if idx == 0 else "",
                        employee.birthday if idx == 0 else "",
                        employee.preferred_faculty if idx == 0 else "",

                        position.department,
                        position.post,
                        position.rate,
                        position.type_of_employment,
                        position.contract_end_date,
                    ])
                
        ExportService.set_width_and_alignment(sheet, [5, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20,])
        temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        return temp_file.name

    async def export_lessons(self) -> str:
        """
        Поля таблицы: "id", "Наименование", "Факультет", "Семестр", "Год", "Компетенции"
        :return: имя файла для экспорта
        """

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Лекциии"

        headers = ["id", "Наименование", "Факультет", "Семестр", "Год", "Поток", "Допустимые Педагоги", "Типы нагрузки"]
        sheet.append(headers)

        ExportService.set_width_and_alignment(sheet, [8, 60, 10, 5, 12, 10, 40, 40])
        async with self.db.session_factory() as session:
            lessons = await session.execute(select(Lesson).distinct())
            lessons = lessons.scalars().unique().all()

            for lesson in lessons:
                types = set()
                for worklaod in lesson.workloads:
                    types.add(worklaod.type)

                sheet.append([
                    lesson.id,
                    lesson.name,
                    lesson.faculty,
                    lesson.semester,
                    lesson.year,
                    lesson.stream,
                    ", ".join([str(name) for name in lesson.employees]) if len(lesson.employees) != 0 else "",
                    ", ".join(types) if len(types) != 0 else "",
                ])

        temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        return temp_file.name

    @staticmethod
    def set_width_and_alignment(sheet: Workbook.active, wights: list[int]) -> None:
        for i, width in enumerate(wights, start=1):
            col_letter = chr(64 + i)
            for cell in sheet[col_letter]:
                cell.alignment = Alignment(horizontal="left")
            sheet.column_dimensions[col_letter].width = width
